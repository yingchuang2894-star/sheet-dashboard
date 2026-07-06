#!/usr/bin/env python3
"""
sync_data.py  ——  下载 Google Sheet → 解析 → 写出 data.js

本地运行（已有 sheet.xlsx）：
    python sync_data.py

本地运行（自动下载）：
    set SPREADSHEET_ID=你的表格ID
    set GOOGLE_SERVICE_ACCOUNT_JSON={"type":"service_account",...}
    python sync_data.py

GitHub Actions 自动运行时也是同样的环境变量。
"""
import json, os, io, sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timezone
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

HERE = Path(__file__).parent
XLSX_PATH = HERE / "sheet.xlsx"
OUTPUT_PATH = HERE / "data.js"

SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID", "")
SA_JSON_STR = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")

# ── 1. 获取 xlsx（优先远程下载，其次用本地文件）────────────────────────────
if SA_JSON_STR and SPREADSHEET_ID:
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload

        creds = service_account.Credentials.from_service_account_info(
            json.loads(SA_JSON_STR),
            scopes=["https://www.googleapis.com/auth/drive.readonly"],
        )
        drive = build("drive", "v3", credentials=creds)
        req = drive.files().export_media(
            fileId=SPREADSHEET_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        XLSX_PATH.write_bytes(buf.getvalue())
        print(f"[sync] 已从 Google Drive 下载 → {XLSX_PATH}")
    except Exception as e:
        print(f"[sync] 下载失败: {e}", file=sys.stderr)
        if not XLSX_PATH.exists():
            sys.exit(1)
        print("[sync] 降级使用本地 sheet.xlsx")
elif XLSX_PATH.exists():
    print(f"[sync] 使用本地文件 {XLSX_PATH}")
else:
    sys.exit(
        "[sync] 错误：找不到 sheet.xlsx，且未设置 SPREADSHEET_ID / GOOGLE_SERVICE_ACCOUNT_JSON"
    )

# ── 2. 解析 xlsx ─────────────────────────────────────────────────────────────
import openpyxl

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)


def fdate(dt):
    if hasattr(dt, "strftime"):
        return dt.strftime("%Y-%m-%d")
    s = str(dt).strip().replace("/", "-")
    parts = s.split("-")
    if len(parts) != 3:
        return None
    y, m, d = parts
    return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"


def cell(ws, r, c):
    return ws.cell(row=r, column=c).value


def r2(v):
    return round(v, 2) if isinstance(v, (int, float)) else v


def r4(v):
    return round(v, 4) if isinstance(v, (int, float)) else v


def sheet_rows(name):
    ws = wb[name]
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if d is not None:
            yield r, d


# 每日数据登记
ws_core = wb["每日数据登记"]
daily_core = {}
for r, date_cell in sheet_rows("每日数据登记"):
    date = fdate(date_cell)
    dau = cell(ws_core, r, 2)
    if not date or not dau:
        continue
    daily_core[date] = {
        "date": date,
        "dau": r2(dau),
        "dau_android": r2(cell(ws_core, r, 11)),
        "dau_h5": r2(cell(ws_core, r, 12)),
        "reg": r2(cell(ws_core, r, 13)),
        "reg_android": r2(cell(ws_core, r, 15)),
        "reg_h5": r2(cell(ws_core, r, 16)),
        "orders": r2(cell(ws_core, r, 19)),
        "orders_success": r2(cell(ws_core, r, 20)),
        "revenue": r2(cell(ws_core, r, 36)),
        "aov": r2(cell(ws_core, r, 35)),
        "pay_success_rate": r4(cell(ws_core, r, 28)),
        "reg_pay_rate": r4(cell(ws_core, r, 26)),
        "retention_pay_rate": r4(cell(ws_core, r, 27)),
        "arpu": r2(cell(ws_core, r, 41)),
        "coin_revenue": r2(cell(ws_core, r, 37)),
        "member_revenue": r2(cell(ws_core, r, 38)),
        "orders_android": r2(cell(ws_core, r, 42)),
        "revenue_android": r2(cell(ws_core, r, 43)),
        "orders_h5": r2(cell(ws_core, r, 46)),
        "revenue_h5": r2(cell(ws_core, r, 47)),
    }

# 数据源（付费率 + 充值档位分层）数值均为实际值÷100，×100还原
ws_src = wb["数据源"]
for r, date_cell in sheet_rows("数据源"):
    date = fdate(date_cell)
    if date in daily_core:
        daily_core[date]["pay_rate"] = r4(cell(ws_src, r, 13))
        for col, key in [(15, "tier_500"), (16, "tier_300"), (17, "tier_200"), (18, "tier_100")]:
            v = cell(ws_src, r, col)
            daily_core[date][key] = round(v * 100, 2) if isinstance(v, (int, float)) else None

# 留存数据
ws_ret = wb["留存数据"]
ret_cols = {
    "android_d1": 2, "h5_d1": 3, "android_d3": 4, "h5_d3": 5,
    "android_d7": 6, "h5_d7": 7, "android_d15": 8, "h5_d15": 9,
    "android_d30": 10, "h5_d30": 11,
    "total_d1": 12, "total_d3": 13, "total_d7": 14, "total_d15": 15, "total_d30": 16,
}
retention = {}
for r, date_cell in sheet_rows("留存数据"):
    date = fdate(date_cell)
    total_d1 = cell(ws_ret, r, 12)
    if not date or total_d1 is None:
        continue
    rec = {}
    for k, c in ret_cols.items():
        v = cell(ws_ret, r, c)
        rec[k] = round(v * 100, 1) if isinstance(v, (int, float)) else None
    retention[date] = rec

# xb月计划（充值流水 + 目标追踪）
ws_plan = wb["xb月计划"]
recharge = {}
for r in range(2, ws_plan.max_row + 1):
    date_cell = cell(ws_plan, r, 1)
    val = cell(ws_plan, r, 2)
    if date_cell is None or val is None:
        continue
    date = fdate(date_cell)
    if date:
        recharge[date] = round(val * 100, 2)

target_keys = {
    "月度目标": "goal", "截止日期": "deadline", "已完成充值": "completed",
    "完成率": "completion_rate", "剩余目标": "remaining", "剩余天数": "days_left",
    "后半月日均需完成": "daily_needed_2h", "理论日均": "theoretical_daily_avg",
    "进度状态": "status", "后半月压力评估": "pressure_assessment",
}
target = {}
for r in range(1, ws_plan.max_row + 1):
    label = cell(ws_plan, r, 4)
    val = cell(ws_plan, r, 5)
    if label is None or label not in target_keys:
        continue
    key = target_keys[label]
    if key == "deadline":
        target[key] = fdate(val) if hasattr(val, "strftime") else val
    elif key in ("goal", "completed", "remaining", "theoretical_daily_avg", "daily_needed_2h"):
        target[key] = round(val * 100, 2) if isinstance(val, (int, float)) else val
    elif key == "completion_rate":
        target[key] = r4(val)
    else:
        target[key] = val

# ── 3. 合并 + 月度聚合 ───────────────────────────────────────────────────────
all_dates = sorted(set(daily_core) | set(retention) | set(recharge))
daily = []
for d in all_dates:
    rec = {"date": d, "month": d[:7]}
    rec.update(daily_core.get(d, {}))
    rec["date"] = d
    for k, v in retention.get(d, {}).items():
        rec[k] = v
    if d in recharge:
        rec["recharge"] = recharge[d]
    daily.append(rec)

months_map = defaultdict(list)
for rec in daily:
    months_map[rec["month"]].append(rec)


def avg(vals, prec=2):
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), prec) if vals else None


def total(vals):
    vals = [v for v in vals if v is not None]
    return round(sum(vals), 2) if vals else None


monthly = []
for m in sorted(months_map):
    recs = months_map[m]
    monthly.append({
        "month": m,
        "days": len(recs),
        "dau_avg": avg([r.get("dau") for r in recs]),
        "reg_total": total([r.get("reg") for r in recs]),
        "orders_total": total([r.get("orders_success") for r in recs]),
        "revenue_total": total([r.get("revenue") for r in recs]),
        "arpu_avg": avg([r.get("arpu") for r in recs]),
        "pay_success_rate_avg": avg([r.get("pay_success_rate") for r in recs], prec=4),
        "total_d1_avg": avg([r.get("total_d1") for r in recs]),
        "total_d7_avg": avg([r.get("total_d7") for r in recs]),
        "total_d30_avg": avg([r.get("total_d30") for r in recs]),
        "recharge_total": total([r.get("recharge") for r in recs]),
        "tier_100_total": total([r.get("tier_100") for r in recs]),
        "tier_200_total": total([r.get("tier_200") for r in recs]),
        "tier_300_total": total([r.get("tier_300") for r in recs]),
        "tier_500_total": total([r.get("tier_500") for r in recs]),
        "coin_total": total([r.get("coin_revenue") for r in recs]),
        "member_total": total([r.get("member_revenue") for r in recs]),
    })

# ── 4. 输出 data.js ──────────────────────────────────────────────────────────
last_date = daily[-1]["date"] if daily else "N/A"
note = (
    f"数据来自 Google Sheets 四个标签页（每日数据登记 / 数据源 / 留存数据 / xb月计划）；"
    f"「留存数据」按÷100缩写填写，已×100还原；"
    f"最后更新：{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}，数据截至 {last_date}。"
)

out = {
    "daily": daily,
    "monthly": monthly,
    "target_may": target,
    "note": note,
}

OUTPUT_PATH.write_text(
    "const DASHBOARD_DATA = " + json.dumps(out, ensure_ascii=False) + ";\n",
    encoding="utf-8",
)

print(f"[sync] 写出 {OUTPUT_PATH}")
print(f"[sync] {len(daily)} 天 | {len(monthly)} 个月 | 截至 {last_date}")
